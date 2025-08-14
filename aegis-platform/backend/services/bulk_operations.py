"""Bulk operations service for import/export of risks, controls, and other entities"""

import io
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Union, Tuple
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_

from config import settings
from models.risk import Risk
from models.assessment import Assessment
from models.framework import Framework, Control
from models.evidence import Evidence
from models.user import User
from models.audit import AuditLog

logger = logging.getLogger(__name__)

class BulkOperationsService:
    """Service for bulk import/export operations"""
    
    def __init__(self):
        self.upload_dir = Path(settings.UPLOADS_DIR) / "bulk_operations"
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Supported formats
        self.supported_formats = ['csv', 'xlsx', 'json']
        
        # Field mappings for different entities
        self.risk_field_mapping = {
            'id': 'id',
            'title': 'title',
            'description': 'description',
            'category': 'category',
            'level': 'level',
            'risk_score': 'risk_score',
            'impact': 'impact',
            'likelihood': 'likelihood',
            'status': 'status',
            'owner': 'owner',
            'due_date': 'due_date',
            'mitigation_strategy': 'mitigation_strategy',
            'residual_risk': 'residual_risk',
            'controls': 'controls',
            'tags': 'tags'
        }
        
        self.control_field_mapping = {
            'id': 'id',
            'control_id': 'control_id',
            'name': 'name',
            'description': 'description',
            'category': 'category',
            'control_type': 'control_type',
            'implementation_status': 'implementation_status',
            'effectiveness': 'effectiveness',
            'test_frequency': 'test_frequency',
            'last_tested': 'last_tested',
            'next_test_date': 'next_test_date',
            'owner': 'owner',
            'evidence_required': 'evidence_required',
            'framework_id': 'framework_id'
        }
    
    def generate_risk_template(self, format_type: str = 'xlsx') -> str:
        """Generate a template file for risk import"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'xlsx':
                filename = f"risk_import_template_{timestamp}.xlsx"
                file_path = self.upload_dir / filename
                
                # Create workbook with template
                wb = Workbook()
                ws = wb.active
                ws.title = "Risk Import Template"
                
                # Headers
                headers = [
                    'title', 'description', 'category', 'level', 'risk_score',
                    'impact', 'likelihood', 'status', 'owner', 'due_date',
                    'mitigation_strategy', 'residual_risk', 'controls', 'tags'
                ]
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                
                # Add example data
                example_data = [
                    [
                        'Data Breach Risk',
                        'Risk of unauthorized access to customer data',
                        'Information Security',
                        'High',
                        8,
                        'High',
                        'Medium',
                        'Open',
                        'john.doe@company.com',
                        '2024-12-31',
                        'Implement multi-factor authentication and encryption',
                        'Medium',
                        'MFA-001,ENC-002',
                        'cybersecurity,compliance'
                    ],
                    [
                        'System Downtime Risk',
                        'Risk of critical system unavailability',
                        'Operational',
                        'Medium',
                        6,
                        'Medium',
                        'Medium',
                        'In Progress',
                        'jane.smith@company.com',
                        '2024-11-30',
                        'Implement redundant systems and backup procedures',
                        'Low',
                        'BAK-001,RED-002',
                        'operations,continuity'
                    ]
                ]
                
                for row_idx, row_data in enumerate(example_data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Add instructions sheet
                ws_instructions = wb.create_sheet("Instructions")
                instructions = [
                    "Risk Import Template Instructions",
                    "",
                    "Required Fields:",
                    "- title: Risk title (required)",
                    "- description: Detailed risk description",
                    "- category: Risk category (e.g., 'Information Security', 'Operational')",
                    "",
                    "Optional Fields:",
                    "- level: Risk level (Critical, High, Medium, Low)",
                    "- risk_score: Numeric risk score (1-10)",
                    "- impact: Impact level (Critical, High, Medium, Low)",
                    "- likelihood: Likelihood level (Very High, High, Medium, Low, Very Low)",
                    "- status: Risk status (Open, In Progress, Resolved, Closed)",
                    "- owner: Risk owner email address",
                    "- due_date: Due date in YYYY-MM-DD format",
                    "- mitigation_strategy: Risk mitigation plan",
                    "- residual_risk: Residual risk level after mitigation",
                    "- controls: Comma-separated list of control IDs",
                    "- tags: Comma-separated list of tags",
                    "",
                    "Notes:",
                    "- Remove the example data before importing your risks",
                    "- Ensure email addresses are valid for risk owners",
                    "- Use consistent formatting for dates (YYYY-MM-DD)",
                    "- Maximum 1000 risks per import"
                ]
                
                for row, instruction in enumerate(instructions, 1):
                    ws_instructions.cell(row=row, column=1, value=instruction)
                
                # Auto-adjust column widths
                for ws in [wb.active, ws_instructions]:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(str(file_path))
                
            elif format_type == 'csv':
                filename = f"risk_import_template_{timestamp}.csv"
                file_path = self.upload_dir / filename
                
                headers = list(self.risk_field_mapping.keys())[1:]  # Exclude ID
                
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(headers)
                    
                    # Add example data
                    writer.writerow([
                        'Data Breach Risk',
                        'Risk of unauthorized access to customer data',
                        'Information Security',
                        'High',
                        8,
                        'High',
                        'Medium',
                        'Open',
                        'john.doe@company.com',
                        '2024-12-31',
                        'Implement multi-factor authentication and encryption',
                        'Medium',
                        'MFA-001,ENC-002',
                        'cybersecurity,compliance'
                    ])
            
            logger.info(f"Risk template generated: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to generate risk template: {e}")
            raise e
    
    def generate_control_template(self, format_type: str = 'xlsx') -> str:
        """Generate a template file for control import"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'xlsx':
                filename = f"control_import_template_{timestamp}.xlsx"
                file_path = self.upload_dir / filename
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Control Import Template"
                
                headers = [
                    'control_id', 'name', 'description', 'category', 'control_type',
                    'implementation_status', 'effectiveness', 'test_frequency',
                    'last_tested', 'next_test_date', 'owner', 'evidence_required',
                    'framework_id'
                ]
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Add example data
                example_data = [
                    [
                        'AC-001',
                        'Multi-Factor Authentication',
                        'Implement MFA for all user accounts',
                        'Access Control',
                        'Preventive',
                        'Implemented',
                        'High',
                        'Quarterly',
                        '2024-01-15',
                        '2024-04-15',
                        'security@company.com',
                        'Yes',
                        1
                    ]
                ]
                
                for row_idx, row_data in enumerate(example_data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(str(file_path))
                
            elif format_type == 'csv':
                filename = f"control_import_template_{timestamp}.csv"
                file_path = self.upload_dir / filename
                
                headers = list(self.control_field_mapping.keys())[1:]  # Exclude ID
                
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(headers)
                    
                    writer.writerow([
                        'AC-001',
                        'Multi-Factor Authentication',
                        'Implement MFA for all user accounts',
                        'Access Control',
                        'Preventive',
                        'Implemented',
                        'High',
                        'Quarterly',
                        '2024-01-15',
                        '2024-04-15',
                        'security@company.com',
                        'Yes',
                        1
                    ])
            
            logger.info(f"Control template generated: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to generate control template: {e}")
            raise e
    
    def parse_import_file(self, file_path: str, entity_type: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse import file and return data with validation errors"""
        try:
            file_path = Path(file_path)
            file_extension = file_path.suffix.lower()
            errors = []
            data = []
            
            if entity_type == 'risks':
                field_mapping = self.risk_field_mapping
            elif entity_type == 'controls':
                field_mapping = self.control_field_mapping
            else:
                raise ValueError(f"Unsupported entity type: {entity_type}")
            
            if file_extension == '.csv':
                data, errors = self._parse_csv_file(file_path, field_mapping)
            elif file_extension in ['.xlsx', '.xls']:
                data, errors = self._parse_excel_file(file_path, field_mapping)
            elif file_extension == '.json':
                data, errors = self._parse_json_file(file_path, field_mapping)
            else:
                errors.append(f"Unsupported file format: {file_extension}")
            
            return data, errors
            
        except Exception as e:
            logger.error(f"Failed to parse import file: {e}")
            return [], [f"Failed to parse file: {str(e)}"]
    
    def _parse_csv_file(self, file_path: Path, field_mapping: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse CSV file"""
        data = []
        errors = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                # Detect delimiter
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                
                for row_num, row in enumerate(reader, 2):  # Start from 2 (after header)
                    try:
                        # Clean and validate row data
                        cleaned_row = {}
                        for field, value in row.items():
                            if field and field.strip() in field_mapping:
                                cleaned_value = value.strip() if isinstance(value, str) else value
                                if cleaned_value:  # Only include non-empty values
                                    cleaned_row[field.strip()] = cleaned_value
                        
                        if cleaned_row:  # Only add non-empty rows
                            data.append(cleaned_row)
                            
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        
        except Exception as e:
            errors.append(f"Failed to read CSV file: {str(e)}")
        
        return data, errors
    
    def _parse_excel_file(self, file_path: Path, field_mapping: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse Excel file"""
        data = []
        errors = []
        
        try:
            # Try to load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Use first worksheet or find the main data sheet
            if len(wb.sheetnames) > 1:
                # Look for sheets that might contain data (not instructions)
                data_sheets = [name for name in wb.sheetnames 
                             if 'instruction' not in name.lower() and 'template' not in name.lower()]
                sheet_name = data_sheets[0] if data_sheets else wb.sheetnames[0]
            else:
                sheet_name = wb.sheetnames[0]
            
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append("")
            
            # Process data rows
            for row_num in range(2, ws.max_row + 1):
                try:
                    row_data = {}
                    has_data = False
                    
                    for col_num, header in enumerate(headers, 1):
                        if header in field_mapping:
                            cell = ws.cell(row=row_num, column=col_num)
                            value = cell.value
                            
                            if value is not None:
                                # Convert to string and clean
                                cleaned_value = str(value).strip() if value else ""
                                if cleaned_value:
                                    row_data[header] = cleaned_value
                                    has_data = True
                    
                    if has_data:
                        data.append(row_data)
                        
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
        except Exception as e:
            errors.append(f"Failed to read Excel file: {str(e)}")
        
        return data, errors
    
    def _parse_json_file(self, file_path: Path, field_mapping: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse JSON file"""
        data = []
        errors = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as jsonfile:
                json_data = json.load(jsonfile)
                
                if isinstance(json_data, list):
                    for item_num, item in enumerate(json_data, 1):
                        try:
                            if isinstance(item, dict):
                                # Filter only mapped fields
                                cleaned_item = {}
                                for field, value in item.items():
                                    if field in field_mapping and value:
                                        cleaned_item[field] = value
                                
                                if cleaned_item:
                                    data.append(cleaned_item)
                            else:
                                errors.append(f"Item {item_num}: Expected object, got {type(item)}")
                                
                        except Exception as e:
                            errors.append(f"Item {item_num}: {str(e)}")
                            
                elif isinstance(json_data, dict):
                    # Single object
                    try:
                        cleaned_item = {}
                        for field, value in json_data.items():
                            if field in field_mapping and value:
                                cleaned_item[field] = value
                        
                        if cleaned_item:
                            data.append(cleaned_item)
                    except Exception as e:
                        errors.append(f"Data: {str(e)}")
                else:
                    errors.append("JSON must contain an object or array of objects")
                    
        except Exception as e:
            errors.append(f"Failed to read JSON file: {str(e)}")
        
        return data, errors
    
    def import_risks(self, 
                    data: List[Dict[str, Any]], 
                    user: User, 
                    db: Session,
                    update_existing: bool = False) -> Dict[str, Any]:
        """Import risks from parsed data"""
        
        results = {
            'total_processed': 0,
            'successful_imports': 0,
            'updated_records': 0,
            'failed_imports': 0,
            'errors': [],
            'imported_ids': []
        }
        
        try:
            for row_num, risk_data in enumerate(data, 1):
                try:
                    # Validate required fields
                    if 'title' not in risk_data or not risk_data['title']:
                        results['errors'].append(f"Row {row_num}: Missing required field 'title'")
                        results['failed_imports'] += 1
                        continue
                    
                    # Check if risk already exists (by title)
                    existing_risk = db.query(Risk).filter(Risk.title == risk_data['title']).first()
                    
                    if existing_risk and not update_existing:
                        results['errors'].append(f"Row {row_num}: Risk '{risk_data['title']}' already exists")
                        results['failed_imports'] += 1
                        continue
                    
                    # Prepare risk data
                    risk_fields = {}
                    
                    # Map fields
                    for field, value in risk_data.items():
                        if field in self.risk_field_mapping:
                            db_field = self.risk_field_mapping[field]
                            
                            # Special handling for specific fields
                            if field == 'risk_score':
                                try:
                                    risk_fields[db_field] = int(float(value))
                                except (ValueError, TypeError):
                                    results['errors'].append(f"Row {row_num}: Invalid risk_score '{value}'")
                                    continue
                            elif field == 'due_date':
                                try:
                                    # Parse date
                                    if isinstance(value, str):
                                        from datetime import datetime
                                        risk_fields[db_field] = datetime.strptime(value, '%Y-%m-%d').date()
                                    else:
                                        risk_fields[db_field] = value
                                except (ValueError, TypeError):
                                    results['errors'].append(f"Row {row_num}: Invalid due_date format '{value}' (use YYYY-MM-DD)")
                                    continue
                            else:
                                risk_fields[db_field] = value
                    
                    # Set created_by
                    risk_fields['created_by'] = user.id
                    
                    if existing_risk and update_existing:
                        # Update existing risk
                        for field, value in risk_fields.items():
                            setattr(existing_risk, field, value)
                        
                        results['updated_records'] += 1
                        results['imported_ids'].append(existing_risk.id)
                        
                        # Log update
                        audit_log = AuditLog(
                            event_type="update",
                            entity_type="risk",
                            entity_id=existing_risk.id,
                            user_id=user.id,
                            action="Risk updated via bulk import",
                            description=f"Risk '{existing_risk.title}' updated via bulk import",
                            source="bulk_import",
                            risk_level="low"
                        )
                        db.add(audit_log)
                        
                    else:
                        # Create new risk
                        new_risk = Risk(**risk_fields)
                        db.add(new_risk)
                        db.flush()  # Get the ID
                        
                        results['successful_imports'] += 1
                        results['imported_ids'].append(new_risk.id)
                        
                        # Log creation
                        audit_log = AuditLog(
                            event_type="create",
                            entity_type="risk",
                            entity_id=new_risk.id,
                            user_id=user.id,
                            action="Risk created via bulk import",
                            description=f"Risk '{new_risk.title}' created via bulk import",
                            source="bulk_import",
                            risk_level="low"
                        )
                        db.add(audit_log)
                    
                    results['total_processed'] += 1
                    
                except Exception as e:
                    logger.error(f"Error importing risk row {row_num}: {e}")
                    results['errors'].append(f"Row {row_num}: {str(e)}")
                    results['failed_imports'] += 1
            
            # Commit all changes
            db.commit()
            
            logger.info(f"Bulk risk import completed: {results['successful_imports']} created, {results['updated_records']} updated, {results['failed_imports']} failed")
            
        except Exception as e:
            db.rollback()
            logger.error(f"Bulk risk import failed: {e}")
            results['errors'].append(f"Import failed: {str(e)}")
        
        return results
    
    def export_risks(self, 
                    db: Session, 
                    format_type: str = 'xlsx',
                    filters: Optional[Dict[str, Any]] = None) -> str:
        """Export risks to file"""
        
        try:
            # Build query with filters
            query = db.query(Risk)
            
            if filters:
                if filters.get('category'):
                    query = query.filter(Risk.category == filters['category'])
                if filters.get('level'):
                    query = query.filter(Risk.level == filters['level'])
                if filters.get('status'):
                    query = query.filter(Risk.status == filters['status'])
                if filters.get('owner'):
                    query = query.filter(Risk.owner == filters['owner'])
            
            risks = query.all()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'xlsx':
                filename = f"risks_export_{timestamp}.xlsx"
                file_path = self.upload_dir / filename
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Risks Export"
                
                # Headers
                headers = [
                    'ID', 'Title', 'Description', 'Category', 'Level', 'Risk Score',
                    'Impact', 'Likelihood', 'Status', 'Owner', 'Due Date',
                    'Mitigation Strategy', 'Residual Risk', 'Created Date', 'Updated Date'
                ]
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Add data
                for row, risk in enumerate(risks, 2):
                    ws.cell(row=row, column=1, value=risk.id)
                    ws.cell(row=row, column=2, value=risk.title)
                    ws.cell(row=row, column=3, value=risk.description)
                    ws.cell(row=row, column=4, value=risk.category)
                    ws.cell(row=row, column=5, value=risk.level)
                    ws.cell(row=row, column=6, value=risk.risk_score)
                    ws.cell(row=row, column=7, value=risk.impact)
                    ws.cell(row=row, column=8, value=risk.likelihood)
                    ws.cell(row=row, column=9, value=risk.status)
                    ws.cell(row=row, column=10, value=risk.owner)
                    ws.cell(row=row, column=11, value=risk.due_date.strftime('%Y-%m-%d') if risk.due_date else '')
                    ws.cell(row=row, column=12, value=risk.mitigation_strategy)
                    ws.cell(row=row, column=13, value=risk.residual_risk)
                    ws.cell(row=row, column=14, value=risk.created_at.strftime('%Y-%m-%d %H:%M:%S') if risk.created_at else '')
                    ws.cell(row=row, column=15, value=risk.updated_at.strftime('%Y-%m-%d %H:%M:%S') if risk.updated_at else '')
                
                wb.save(str(file_path))
                
            elif format_type == 'csv':
                filename = f"risks_export_{timestamp}.csv"
                file_path = self.upload_dir / filename
                
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    
                    # Headers
                    writer.writerow([
                        'ID', 'Title', 'Description', 'Category', 'Level', 'Risk Score',
                        'Impact', 'Likelihood', 'Status', 'Owner', 'Due Date',
                        'Mitigation Strategy', 'Residual Risk', 'Created Date', 'Updated Date'
                    ])
                    
                    # Data
                    for risk in risks:
                        writer.writerow([
                            risk.id,
                            risk.title,
                            risk.description,
                            risk.category,
                            risk.level,
                            risk.risk_score,
                            risk.impact,
                            risk.likelihood,
                            risk.status,
                            risk.owner,
                            risk.due_date.strftime('%Y-%m-%d') if risk.due_date else '',
                            risk.mitigation_strategy,
                            risk.residual_risk,
                            risk.created_at.strftime('%Y-%m-%d %H:%M:%S') if risk.created_at else '',
                            risk.updated_at.strftime('%Y-%m-%d %H:%M:%S') if risk.updated_at else ''
                        ])
            
            elif format_type == 'json':
                filename = f"risks_export_{timestamp}.json"
                file_path = self.upload_dir / filename
                
                risks_data = []
                for risk in risks:
                    risk_dict = {
                        'id': risk.id,
                        'title': risk.title,
                        'description': risk.description,
                        'category': risk.category,
                        'level': risk.level,
                        'risk_score': risk.risk_score,
                        'impact': risk.impact,
                        'likelihood': risk.likelihood,
                        'status': risk.status,
                        'owner': risk.owner,
                        'due_date': risk.due_date.strftime('%Y-%m-%d') if risk.due_date else None,
                        'mitigation_strategy': risk.mitigation_strategy,
                        'residual_risk': risk.residual_risk,
                        'created_at': risk.created_at.isoformat() if risk.created_at else None,
                        'updated_at': risk.updated_at.isoformat() if risk.updated_at else None
                    }
                    risks_data.append(risk_dict)
                
                with open(file_path, 'w', encoding='utf-8') as jsonfile:
                    json.dump(risks_data, jsonfile, indent=2, ensure_ascii=False)
            
            logger.info(f"Risks export completed: {len(risks)} risks exported to {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export risks: {e}")
            raise e
    
    def validate_import_data(self, data: List[Dict[str, Any]], entity_type: str) -> List[str]:
        """Validate import data and return list of errors"""
        errors = []
        
        if entity_type == 'risks':
            required_fields = ['title']
            valid_levels = ['Critical', 'High', 'Medium', 'Low']
            valid_statuses = ['Open', 'In Progress', 'Resolved', 'Closed']
            
            for row_num, item in enumerate(data, 1):
                # Check required fields
                for field in required_fields:
                    if field not in item or not item[field]:
                        errors.append(f"Row {row_num}: Missing required field '{field}'")
                
                # Validate level
                if 'level' in item and item['level'] not in valid_levels:
                    errors.append(f"Row {row_num}: Invalid level '{item['level']}'. Must be one of: {valid_levels}")
                
                # Validate status
                if 'status' in item and item['status'] not in valid_statuses:
                    errors.append(f"Row {row_num}: Invalid status '{item['status']}'. Must be one of: {valid_statuses}")
                
                # Validate risk score
                if 'risk_score' in item:
                    try:
                        score = float(item['risk_score'])
                        if score < 0 or score > 10:
                            errors.append(f"Row {row_num}: Risk score must be between 0 and 10")
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Risk score must be a number")
        
        return errors

# Global service instance
bulk_operations_service = BulkOperationsService()