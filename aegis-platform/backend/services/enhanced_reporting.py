"""Enhanced PDF and Excel reporting service with branded templates"""

import io
import os
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
import base64

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from weasyprint import HTML, CSS

# Excel Generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as OpenpyxlImage

# Template Engine
from jinja2 import Template, Environment, FileSystemLoader

from config import settings

logger = logging.getLogger(__name__)

class BrandingConfig:
    """Configuration for branded reports"""
    
    PRIMARY_COLOR = colors.HexColor('#6366f1')  # Indigo
    SECONDARY_COLOR = colors.HexColor('#8b5cf6')  # Purple
    ACCENT_COLOR = colors.HexColor('#ec4899')  # Pink
    SUCCESS_COLOR = colors.HexColor('#10b981')  # Green
    WARNING_COLOR = colors.HexColor('#f59e0b')  # Amber
    DANGER_COLOR = colors.HexColor('#ef4444')   # Red
    
    LOGO_PATH = None  # Path to company logo
    COMPANY_NAME = "Aegis Risk Management"
    TAGLINE = "Enterprise Security & Compliance Platform"
    WEBSITE = "https://aegis-platform.com"
    
    # Fonts
    FONT_TITLE = "Helvetica-Bold"
    FONT_HEADER = "Helvetica-Bold"
    FONT_BODY = "Helvetica"
    FONT_SMALL = "Helvetica"

class EnhancedReportingService:
    """Enhanced reporting service with PDF/Excel generation and branding"""
    
    def __init__(self):
        self.reports_dir = Path(settings.UPLOADS_DIR) / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
        self.templates_dir = Path(__file__).parent / "report_templates"
        self.templates_dir.mkdir(exist_ok=True)
        
        # Initialize Jinja2 environment
        self.jinja_env = Environment(
            loader=FileSystemLoader(str(self.templates_dir))
        )
        
        # Custom styles for ReportLab
        self.styles = self._create_custom_styles()
        
        # Branding configuration
        self.branding = BrandingConfig()
    
    def _create_custom_styles(self) -> Dict[str, ParagraphStyle]:
        """Create custom paragraph styles for branded reports"""
        styles = getSampleStyleSheet()
        
        custom_styles = {
            'CustomTitle': ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=24,
                textColor=BrandingConfig.PRIMARY_COLOR,
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName=BrandingConfig.FONT_TITLE
            ),
            'CustomHeading1': ParagraphStyle(
                'CustomHeading1',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=BrandingConfig.PRIMARY_COLOR,
                spaceBefore=20,
                spaceAfter=12,
                fontName=BrandingConfig.FONT_HEADER
            ),
            'CustomHeading2': ParagraphStyle(
                'CustomHeading2',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=BrandingConfig.SECONDARY_COLOR,
                spaceBefore=16,
                spaceAfter=8,
                fontName=BrandingConfig.FONT_HEADER
            ),
            'CustomBody': ParagraphStyle(
                'CustomBody',
                parent=styles['Normal'],
                fontSize=10,
                spaceBefore=6,
                spaceAfter=6,
                fontName=BrandingConfig.FONT_BODY
            ),
            'CustomMetric': ParagraphStyle(
                'CustomMetric',
                parent=styles['Normal'],
                fontSize=12,
                textColor=BrandingConfig.PRIMARY_COLOR,
                alignment=TA_CENTER,
                fontName=BrandingConfig.FONT_HEADER
            )
        }
        
        return custom_styles
    
    def _create_header_footer(self, canvas, doc):
        """Create branded header and footer for PDF pages"""
        canvas.saveState()
        
        # Header
        canvas.setFillColor(BrandingConfig.PRIMARY_COLOR)
        canvas.rect(0, doc.height + doc.topMargin - 40, doc.width + doc.leftMargin + doc.rightMargin, 40, fill=1)
        
        canvas.setFillColor(colors.white)
        canvas.setFont(BrandingConfig.FONT_TITLE, 16)
        canvas.drawString(doc.leftMargin, doc.height + doc.topMargin - 25, BrandingConfig.COMPANY_NAME)
        
        canvas.setFont(BrandingConfig.FONT_SMALL, 8)
        canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.topMargin - 25, BrandingConfig.TAGLINE)
        
        # Footer
        canvas.setFillColor(BrandingConfig.PRIMARY_COLOR)
        canvas.setFont(BrandingConfig.FONT_SMALL, 8)
        canvas.drawString(doc.leftMargin, 30, f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        canvas.drawRightString(doc.width + doc.leftMargin, 30, f"Page {doc.page}")
        canvas.drawCentredString(doc.width/2 + doc.leftMargin, 30, BrandingConfig.WEBSITE)
        
        canvas.restoreState()
    
    def _create_risk_chart(self, risk_data: Dict[str, int]) -> Drawing:
        """Create a pie chart for risk distribution"""
        drawing = Drawing(300, 200)
        
        pie = Pie()
        pie.x = 50
        pie.y = 50
        pie.width = 200
        pie.height = 200
        
        pie.data = list(risk_data.values())
        pie.labels = list(risk_data.keys())
        pie.slices.strokeWidth = 0.5
        
        # Color mapping for risk levels
        colors_map = {
            'Critical': BrandingConfig.DANGER_COLOR,
            'High': BrandingConfig.WARNING_COLOR,
            'Medium': BrandingConfig.SECONDARY_COLOR,
            'Low': BrandingConfig.SUCCESS_COLOR
        }
        
        for i, label in enumerate(pie.labels):
            pie.slices[i].fillColor = colors_map.get(label, BrandingConfig.PRIMARY_COLOR)
        
        drawing.add(pie)
        return drawing
    
    async def generate_branded_risk_register_pdf(self, risks_data: List[Dict[str, Any]], 
                                               organization: str = None) -> str:
        """Generate branded PDF risk register report"""
        try:
            # Prepare filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            pdf_filename = f"risk_register_branded_{timestamp}.pdf"
            pdf_path = self.reports_dir / pdf_filename
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=100,
                bottomMargin=72
            )
            
            # Build document content
            content = []
            
            # Title
            title = Paragraph(
                f"Risk Register Report<br/><font size=12>{organization or BrandingConfig.COMPANY_NAME}</font>",
                self.styles['CustomTitle']
            )
            content.append(title)
            content.append(Spacer(1, 20))
            
            # Summary metrics
            risk_summary = {
                'Critical': len([r for r in risks_data if r.get('level', '').lower() == 'critical']),
                'High': len([r for r in risks_data if r.get('level', '').lower() == 'high']),
                'Medium': len([r for r in risks_data if r.get('level', '').lower() == 'medium']),
                'Low': len([r for r in risks_data if r.get('level', '').lower() == 'low'])
            }
            
            # Summary table
            summary_data = [
                ['Risk Level', 'Count', 'Percentage'],
                ['Critical', str(risk_summary['Critical']), f"{(risk_summary['Critical']/len(risks_data)*100):.1f}%" if risks_data else "0%"],
                ['High', str(risk_summary['High']), f"{(risk_summary['High']/len(risks_data)*100):.1f}%" if risks_data else "0%"],
                ['Medium', str(risk_summary['Medium']), f"{(risk_summary['Medium']/len(risks_data)*100):.1f}%" if risks_data else "0%"],
                ['Low', str(risk_summary['Low']), f"{(risk_summary['Low']/len(risks_data)*100):.1f}%" if risks_data else "0%"]
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 1*inch, 1.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BrandingConfig.PRIMARY_COLOR),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), BrandingConfig.FONT_HEADER),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(Paragraph("Executive Summary", self.styles['CustomHeading1']))
            content.append(summary_table)
            content.append(Spacer(1, 20))
            
            # Risk distribution chart
            if any(risk_summary.values()):
                content.append(Paragraph("Risk Distribution", self.styles['CustomHeading2']))
                risk_chart = self._create_risk_chart(risk_summary)
                content.append(risk_chart)
                content.append(Spacer(1, 20))
            
            # Detailed risk table
            content.append(Paragraph("Detailed Risk Register", self.styles['CustomHeading1']))
            
            if risks_data:
                risk_table_data = [
                    ['ID', 'Title', 'Category', 'Level', 'Score', 'Status', 'Owner']
                ]
                
                for risk in risks_data[:50]:  # Limit to first 50 risks for PDF
                    risk_table_data.append([
                        str(risk.get('id', 'N/A')),
                        risk.get('title', 'N/A')[:30] + ('...' if len(risk.get('title', '')) > 30 else ''),
                        risk.get('category', 'N/A'),
                        risk.get('level', 'N/A'),
                        str(risk.get('score', 'N/A')),
                        risk.get('status', 'N/A'),
                        risk.get('owner', 'N/A')[:15] + ('...' if len(risk.get('owner', '')) > 15 else '')
                    ])
                
                risk_table = Table(risk_table_data, colWidths=[0.5*inch, 2*inch, 1*inch, 0.8*inch, 0.6*inch, 0.8*inch, 1*inch])
                risk_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), BrandingConfig.PRIMARY_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), BrandingConfig.FONT_HEADER),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                
                content.append(risk_table)
            else:
                content.append(Paragraph("No risks found in the system.", self.styles['CustomBody']))
            
            # Build PDF with custom header/footer
            doc.build(content, onFirstPage=self._create_header_footer, onLaterPages=self._create_header_footer)
            
            logger.info(f"Branded risk register PDF generated: {pdf_path}")
            return str(pdf_path)
            
        except Exception as e:
            logger.error(f"Failed to generate branded risk register PDF: {e}")
            raise e
    
    async def generate_excel_risk_register(self, risks_data: List[Dict[str, Any]], 
                                         organization: str = None) -> str:
        """Generate Excel risk register with charts and formatting"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"risk_register_{timestamp}.xlsx"
            excel_path = self.reports_dir / excel_filename
            
            # Create workbook
            wb = Workbook()
            
            # Remove default worksheet
            wb.remove(wb.active)
            
            # Create Summary worksheet
            ws_summary = wb.create_sheet("Summary")
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Summary sheet content
            ws_summary['A1'] = f"Risk Register Summary - {organization or BrandingConfig.COMPANY_NAME}"
            ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
            ws_summary.merge_cells('A1:F1')
            
            ws_summary['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary['A4'] = f"Total Risks: {len(risks_data)}"
            
            # Risk level summary
            risk_summary = {
                'Critical': len([r for r in risks_data if r.get('level', '').lower() == 'critical']),
                'High': len([r for r in risks_data if r.get('level', '').lower() == 'high']),
                'Medium': len([r for r in risks_data if r.get('level', '').lower() == 'medium']),
                'Low': len([r for r in risks_data if r.get('level', '').lower() == 'low'])
            }
            
            # Summary table
            ws_summary['A6'] = "Risk Level"
            ws_summary['B6'] = "Count"
            ws_summary['C6'] = "Percentage"
            
            for col in ['A6', 'B6', 'C6']:
                ws_summary[col].font = header_font
                ws_summary[col].fill = header_fill
                ws_summary[col].border = border
            
            row = 7
            for level, count in risk_summary.items():
                ws_summary[f'A{row}'] = level
                ws_summary[f'B{row}'] = count
                ws_summary[f'C{row}'] = f"{(count/len(risks_data)*100):.1f}%" if risks_data else "0%"
                
                # Color coding for risk levels
                color_map = {
                    'Critical': "EF4444",
                    'High': "F59E0B", 
                    'Medium': "8B5CF6",
                    'Low': "10B981"
                }
                
                for col in ['A', 'B', 'C']:
                    ws_summary[f'{col}{row}'].border = border
                    if level in color_map:
                        ws_summary[f'A{row}'].fill = PatternFill(start_color=color_map[level], 
                                                               end_color=color_map[level], fill_type="solid")
                        ws_summary[f'A{row}'].font = Font(color="FFFFFF", bold=True)
                
                row += 1
            
            # Create pie chart for risk distribution
            if any(risk_summary.values()):
                chart = PieChart()
                chart.title = "Risk Distribution by Level"
                
                data = Reference(ws_summary, min_col=2, min_row=6, max_row=10, max_col=2)
                categories = Reference(ws_summary, min_col=1, min_row=7, max_row=10)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                ws_summary.add_chart(chart, "E6")
            
            # Create detailed risks worksheet
            ws_risks = wb.create_sheet("Risk Details")
            
            # Headers for detailed risk data
            headers = ['ID', 'Title', 'Description', 'Category', 'Level', 'Score', 
                      'Status', 'Owner', 'Created Date', 'Due Date', 'Mitigation']
            
            for col, header in enumerate(headers, 1):
                cell = ws_risks.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Add risk data
            for row, risk in enumerate(risks_data, 2):
                ws_risks.cell(row=row, column=1, value=risk.get('id', ''))
                ws_risks.cell(row=row, column=2, value=risk.get('title', ''))
                ws_risks.cell(row=row, column=3, value=risk.get('description', ''))
                ws_risks.cell(row=row, column=4, value=risk.get('category', ''))
                ws_risks.cell(row=row, column=5, value=risk.get('level', ''))
                ws_risks.cell(row=row, column=6, value=risk.get('score', ''))
                ws_risks.cell(row=row, column=7, value=risk.get('status', ''))
                ws_risks.cell(row=row, column=8, value=risk.get('owner', ''))
                ws_risks.cell(row=row, column=9, value=risk.get('created_date', ''))
                ws_risks.cell(row=row, column=10, value=risk.get('due_date', ''))
                ws_risks.cell(row=row, column=11, value=risk.get('mitigation', ''))
                
                # Color coding for risk level
                level = risk.get('level', '').lower()
                if level in ['critical', 'high', 'medium', 'low']:
                    color_map = {
                        'critical': "FFEBEE",
                        'high': "FFF3E0", 
                        'medium': "F3E5F5",
                        'low': "E8F5E8"
                    }
                    fill = PatternFill(start_color=color_map[level], end_color=color_map[level], fill_type="solid")
                    
                    for col in range(1, len(headers) + 1):
                        ws_risks.cell(row=row, column=col).fill = fill
                        ws_risks.cell(row=row, column=col).border = border
            
            # Auto-adjust column widths
            for ws in [ws_summary, ws_risks]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(str(excel_path))
            
            logger.info(f"Excel risk register generated: {excel_path}")
            return str(excel_path)
            
        except Exception as e:
            logger.error(f"Failed to generate Excel risk register: {e}")
            raise e
    
    async def generate_executive_dashboard_pdf(self, dashboard_data: Dict[str, Any]) -> str:
        """Generate executive dashboard PDF with metrics and charts"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            pdf_filename = f"executive_dashboard_{timestamp}.pdf"
            pdf_path = self.reports_dir / pdf_filename
            
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=100,
                bottomMargin=72
            )
            
            content = []
            
            # Title
            title = Paragraph("Executive Security Dashboard", self.styles['CustomTitle'])
            content.append(title)
            content.append(Spacer(1, 20))
            
            # Key metrics in a grid
            metrics_data = [
                ['Metric', 'Current Value', 'Status'],
                ['Total Active Risks', str(dashboard_data.get('total_risks', 0)), '📊'],
                ['Critical Risks', str(dashboard_data.get('critical_risks', 0)), '🚨'],
                ['Compliance Score', f"{dashboard_data.get('compliance_score', 0)}%", '✅'],
                ['Open Tasks', str(dashboard_data.get('open_tasks', 0)), '📋'],
                ['Recent Assessments', str(dashboard_data.get('recent_assessments', 0)), '🔍']
            ]
            
            metrics_table = Table(metrics_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BrandingConfig.PRIMARY_COLOR),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), BrandingConfig.FONT_HEADER),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            content.append(Paragraph("Key Security Metrics", self.styles['CustomHeading1']))
            content.append(metrics_table)
            content.append(Spacer(1, 30))
            
            # Top risks section
            content.append(Paragraph("Top Risks Requiring Attention", self.styles['CustomHeading1']))
            
            top_risks = dashboard_data.get('top_risks', [])
            if top_risks:
                for i, risk in enumerate(top_risks[:5], 1):
                    risk_text = f"{i}. <b>{risk.get('title', 'Unknown Risk')}</b><br/>"
                    risk_text += f"   Level: {risk.get('level', 'Unknown')} | "
                    risk_text += f"Score: {risk.get('score', 'N/A')} | "
                    risk_text += f"Owner: {risk.get('owner', 'Unassigned')}<br/><br/>"
                    
                    content.append(Paragraph(risk_text, self.styles['CustomBody']))
            else:
                content.append(Paragraph("No critical risks identified.", self.styles['CustomBody']))
            
            # Build PDF
            doc.build(content, onFirstPage=self._create_header_footer, onLaterPages=self._create_header_footer)
            
            logger.info(f"Executive dashboard PDF generated: {pdf_path}")
            return str(pdf_path)
            
        except Exception as e:
            logger.error(f"Failed to generate executive dashboard PDF: {e}")
            raise e
    
    async def generate_compliance_report_excel(self, compliance_data: Dict[str, Any]) -> str:
        """Generate comprehensive compliance report in Excel format"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"compliance_report_{timestamp}.xlsx"
            excel_path = self.reports_dir / excel_filename
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            
            # Overview sheet
            ws_overview = wb.create_sheet("Compliance Overview")
            
            ws_overview['A1'] = "Compliance Status Report"
            ws_overview['A1'].font = Font(bold=True, size=16, color="6366F1")
            ws_overview.merge_cells('A1:D1')
            
            ws_overview['A3'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d')}"
            ws_overview['A4'] = f"Overall Compliance Score: {compliance_data.get('overall_score', 0)}%"
            
            # Framework compliance by category
            frameworks = compliance_data.get('frameworks', [])
            
            if frameworks:
                ws_overview['A6'] = "Framework Compliance Summary"
                ws_overview['A6'].font = Font(bold=True, size=14)
                
                headers = ['Framework', 'Total Controls', 'Implemented', 'Compliance %', 'Status']
                for col, header in enumerate(headers, 1):
                    cell = ws_overview.cell(row=8, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                for row, framework in enumerate(frameworks, 9):
                    ws_overview.cell(row=row, column=1, value=framework.get('name', ''))
                    ws_overview.cell(row=row, column=2, value=framework.get('total_controls', 0))
                    ws_overview.cell(row=row, column=3, value=framework.get('implemented_controls', 0))
                    ws_overview.cell(row=row, column=4, value=f"{framework.get('compliance_percentage', 0)}%")
                    
                    # Status based on compliance percentage
                    compliance_pct = framework.get('compliance_percentage', 0)
                    if compliance_pct >= 90:
                        status = "Excellent"
                        color = "10B981"
                    elif compliance_pct >= 75:
                        status = "Good"
                        color = "F59E0B"
                    elif compliance_pct >= 50:
                        status = "Needs Improvement"
                        color = "EF4444"
                    else:
                        status = "Critical"
                        color = "DC2626"
                    
                    status_cell = ws_overview.cell(row=row, column=5, value=status)
                    status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    status_cell.font = Font(color="FFFFFF", bold=True)
            
            # Add detailed sheets for each framework
            for framework in frameworks:
                sheet_name = framework.get('name', 'Framework')[:31]  # Excel sheet name limit
                ws_detail = wb.create_sheet(sheet_name)
                
                ws_detail['A1'] = f"{framework.get('name', 'Framework')} - Detailed Controls"
                ws_detail['A1'].font = Font(bold=True, size=14, color="6366F1")
                ws_detail.merge_cells('A1:F1')
                
                # Control details headers
                control_headers = ['Control ID', 'Control Name', 'Category', 'Status', 'Evidence Count', 'Comments']
                for col, header in enumerate(control_headers, 1):
                    cell = ws_detail.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Add control data
                controls = framework.get('controls', [])
                for row, control in enumerate(controls, 4):
                    ws_detail.cell(row=row, column=1, value=control.get('control_id', ''))
                    ws_detail.cell(row=row, column=2, value=control.get('name', ''))
                    ws_detail.cell(row=row, column=3, value=control.get('category', ''))
                    ws_detail.cell(row=row, column=4, value=control.get('status', ''))
                    ws_detail.cell(row=row, column=5, value=control.get('evidence_count', 0))
                    ws_detail.cell(row=row, column=6, value=control.get('comments', ''))
                    
                    # Color coding based on status
                    status = control.get('status', '').lower()
                    if 'implemented' in status:
                        fill_color = "E8F5E8"
                    elif 'partial' in status:
                        fill_color = "FFF3E0"
                    else:
                        fill_color = "FFEBEE"
                    
                    for col in range(1, len(control_headers) + 1):
                        ws_detail.cell(row=row, column=col).fill = PatternFill(
                            start_color=fill_color, end_color=fill_color, fill_type="solid"
                        )
            
            # Auto-adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(str(excel_path))
            
            logger.info(f"Compliance report Excel generated: {excel_path}")
            return str(excel_path)
            
        except Exception as e:
            logger.error(f"Failed to generate compliance report Excel: {e}")
            raise e

# Global service instance
enhanced_reporting_service = EnhancedReportingService()